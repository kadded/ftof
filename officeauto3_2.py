# -*- coding: utf-8 -*-
#############################################################################################
# Дополнительные функции ver. 3.2
#############################################################################################

#############################################################################################
# Подключение модулей (не требует редактирования)
#############################################################################################
import sys
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell, Cell
# from openpyxl.utils import get_column_letter, column_index_from_string,
from openpyxl.utils.cell import get_column_letter, column_index_from_string, coordinate_from_string
from openpyxl.cell.cell import TYPE_STRING, TYPE_NUMERIC, TYPE_BOOL, TYPE_NULL, TYPE_FORMULA, TYPE_ERROR
from decimal import Decimal
#import pandas as pd
import copy
from copy import deepcopy
from openpyxl.formula.translate import Translator



#############################################################################################
# Вспомогательные функции (не требует редактирования)
#############################################################################################
def value(cell, sheet:Worksheet= False):
    if sheet:
        if isinstance(cell, str):
            cell = sheet[cell]
    cell_type = cell.data_type

    if cell_type == TYPE_NUMERIC:
        # Проверка на дату
        try:
            if cell.is_date:
                return cell.value.date()
            else:
                pass
        except:
            return ''
    val = str(cell.value)
    if cell_type == TYPE_NULL:
        return '' # None лучше вернуть пустую строку или строку 'None'
    try:
        float(val)
        return Decimal(val)
    except:
        try:
            return int(val)
        except:
            return val

    cell_type = cell.data_type

    if cell_type == TYPE_STRING:
        return str(cell.value)
    elif cell_type == TYPE_BOOL:
        return str(bool(cell.value))
    elif cell_type == TYPE_NULL:
        return '' # None лучше вернуть пустую строку или строку 'None'
    elif cell_type == TYPE_FORMULA:
        # Если ячейка содержит формулу, возвращаем результат вычисления формулы
        return cell.value
    elif cell_type == TYPE_ERROR:
        # Если ячейка содержит ошибку, возвращаем код ошибки
        return cell.value
    else:
        # Если тип данных неизвестен, возвращаем значение ячейки как есть
        return cell.value




def copy_cell(cell: Cell, newRow: int = False, newCol: int=False, targetCell: Cell=None) -> None:
    """копирует ячейку в другую ячейку. Для входа используется именна объект ячейки
    """
    if targetCell:
        newRow = targetCell.row
        newCol = targetCell.column
        new_cell = targetCell
    else:
        new_cell = cell.parent.cell(row=newRow, column=newCol)
    newColL = get_column_letter(newCol)
    curcolLetter = get_column_letter(cell.column)
    curRow=cell.row

    if cell.data_type == TYPE_FORMULA:
        new_cell.value = Translator(copy.copy(cell.value), f"{curcolLetter}{curRow}").translate_formula(f"{newColL}{newRow}")
    else:
        new_cell.value = copy.copy(cell.value)
    new_cell.data_type = copy.copy(cell.data_type)
    if cell.has_style:
        new_cell.font = copy.copy(cell.font)
        new_cell.border = copy.copy(cell.border)
        new_cell.fill = copy.copy(cell.fill)
        new_cell.number_format = copy.copy(cell.number_format)
        new_cell.protection = copy.copy(cell.protection)
        new_cell.alignment = copy.copy(cell.alignment)





def get_cell_info(cell: Cell, sheet: Worksheet, mergedCellsDict: dict):
    '''Функйия-метод для показа свойств ячейки. Возвращает набор переменных
    это_Коренная, это_Объединенная, Коренная, ЗначениеЯчейки :

        это_Коренная - True если передана коренная ячейка какого-то диапазона
        это_Объединенная - True если ячейка объединена в каком-то диапазоне, но не коренная
        Коренная - возвращает коренную ячейку диапазона, в которую входит переданная объединенная ячейка
        ЗначениеЯчейки - возвращает значение ячейки или значение её коренной ячейки, если она в диапазоне
                с помощью функции value() из библиотечки автоматизации.
        ПОМНИ! У объедененной ячейки НЕТ своего значения. Всегда получаем его от коренно ячейки диапазона
    '''
    if isinstance(cell, str):
        cell=sheet[cell]
    if not isinstance(cell, MergedCell):  # Ячейка не является объединенной
        val = value(cell)
        if not cell.coordinate in mergedCellsDict['ключиОтсортированные']: # Ячейка ещё и не является коренной
            return False, False, False, val # это_Коренная? это_Объединенная? Коренная, значение -> Обычная ячейка одинокая
        else:
            return True,  False, cell, val # это_Коренная? это_Объединенная? Коренная, значение -> Коренная ячейка и её же вернуть как коренную

    column, row = cell.column, cell.row
    for rootCellCoord in mergedCellsDict['ключиОтсортированные']:
        if (column_index_from_string(coordinate_from_string(rootCellCoord)[0]) >= column) and (coordinate_from_string(rootCellCoord)[1] >= row):
            raise Exception('Ошибка в логике') # Эта строка никогда не должна отработать
        if mergedCellsDict['ячейки'][rootCellCoord]['всеЯчейкиВнутри'].__contains__(cell.coordinate):
            # mergedCellsDict['ячейки'][rootCellCoord]['всеЯчейкиВнутри'].remove(cell) # Нельзя удалять - могут понадобиться еще раз
            # if not mergedCellsDict['ячейки'][rootCellCoord]['всеЯчейкиВнутри']:
            #     del mergedCellsDict['ячейки'][rootCellCoord]
            rootCell = sheet[rootCellCoord]
            val = value(rootCell)
            return False, True, rootCell, val # это_Коренная? это_Объединенная? Коренная ЗначениеЯчейки -> объединенная ячейка и у ней есть коренная



def multipleRows(row:dict, multer=int)->int:
    '''Функция или метод для тиражирования строк. На вход подаем строку в виде готового словаря со всеми
    колонками. На выходе получаем список из словарей.'''
    return list([row.copy() for ind in range(multer)])

def fillColumns(rows:list, key: str, value):
    '''Функция или метод - в каждом объекте входящего списка даёт ключу 'key' значение 'value'.
    Можно использовать, например,  после применения функции multipleRows()'''
    for ind in range(len(rows)):
        rows[ind][key] = value
    return rows





def get_grouped_rows(колонкаГруппировки, словарьОбъединенныхЯчеек, номерПервойСтроки, номерПоследнейСтроки)->list:
    '''Возвращает список списков, с группировкой номеров строк по группирующей колонке'''
    if колонкаГруппировки:
        сгруппированныеНомераСтрок = []
        justGrouppedRows = 0
        for номерТекущейСтроки in range(номерПервойСтроки, номерПоследнейСтроки + 1):
            if номерТекущейСтроки <= justGrouppedRows:
                continue
            адресЯчейкиГруппировки = f'{колонкаГруппировки}{номерТекущейСтроки}'

            if адресЯчейкиГруппировки in словарьОбъединенныхЯчеек['ячейки']:
                сгруппированныеНомераСтрок.append( [r[0][0] for r in словарьОбъединенныхЯчеек['ячейки'][адресЯчейкиГруппировки]['всеЯчейкиВнутри'].rows])
                justGrouppedRows = сгруппированныеНомераСтрок[-1][-1]
            else:
                сгруппированныеНомераСтрок.append([номерТекущейСтроки])
                justGrouppedRows = номерТекущейСтроки
    else:
        сгруппированныеНомераСтрок = range([номерПервойСтроки], номерПоследнейСтроки + 1)
    return сгруппированныеНомераСтрок


class СцепЯчейки:
    """Класс для работы с объединенными ячейками"""
    def __init__(self, sheet:Worksheet):
        merged_cells__ranges = sheet.merged_cells.ranges
        self.sheet = sheet
        self.словарьОбъединенныхЯчеек = {
            'ячейки': {range.start_cell.coordinate: {'всеЯчейкиВнутри': range, 'значение': []} for range
                       in merged_cells__ranges}}
        self.словарьОбъединенныхЯчеек['ключиОтсортированные'] = sorted(self.словарьОбъединенныхЯчеек['ячейки'].keys(),
                                                                  key=lambda x: coordinate_from_string(x))

    def get_cell_info(self, cell:Cell):
        '''Функция-метод для показа свойств ячейки. Возвращает набор переменных
        это_Коренная, это_Объединенная, Коренная, ЗначениеЯчейки :

            это_Коренная - True если передана коренная ячейка какого-то диапазона
            это_Объединенная - True если ячейка объединена в каком-то диапазоне, но не коренная
            Коренная - возвращает коренную ячейку диапазона, в которую входит переданная объединенная ячейка
            ЗначениеЯчейки - возвращает значение ячейки с помощью функции value() из библиотечки автоматизации
        '''
        return get_cell_info(cell= cell, sheet=self.sheet, mergedCellsDict=self.словарьОбъединенныхЯчеек)


class ЛистПроцессор():  #
    """Класс для работы с листом"""
    def __init__(self, sheet:Worksheet):
        сцепЯчейки =СцепЯчейки(sheet)
        self.sheet = sheet
        self.rowSliceCounter = 0
        self. numberRows = 0
        self.последняяКолонка = sheet.max_column
        self.букваПоследнейКолонки = get_column_letter(sheet.max_column)
        self.словарьОбъединенныхЯчеек = сцепЯчейки.словарьОбъединенныхЯчеек

    def copy_and_add_row_TO(self, rowNumber:int, targetSheet:Worksheet):
        """Копирование строки и вставка на [другой/этот] лист.  При этом учитываем перенос адресов ячеек.
        Но при переносе ссылки на новый Лист"""

        sourceRow=self.sheet[rowNumber]
        targetRowNumber = targetSheet.max_row + 1
        # targetSheet.insert_rows(targetRowNumber, 1)
        СтильСтроки = self.sheet.row_dimensions[rowNumber]
        targetSheet.row_dimensions[rowNumber] = copy.copy(СтильСтроки)
        for cell in sourceRow:
            cellCol = cell.column
            targetCell = targetSheet.cell(targetRowNumber, cellCol)
            copy_cell(cell, targetCell=targetCell)

    def перенестиТитул_на(self, номераСтрокТитула:[], targetSheet:Worksheet):
        '''Перенос Титула на другой лист'''

        if not len(номераСтрокТитула):
            return
        for номерСтроки in номераСтрокТитула:
            # перенесем строки
            sourceRow = self.sheet[номерСтроки]
            targetRowNumber = номерСтроки
            СтильСтроки = self.sheet.row_dimensions[номерСтроки]
            targetSheet.row_dimensions[targetRowNumber] = copy.copy(СтильСтроки)
            for cell in sourceRow:
                cellCol = cell.column
                targetCell = targetSheet.cell(targetRowNumber, cellCol)
                copy_cell(cell, targetCell=targetCell)
        # Перенос стиля столбцов

        for stN in range(1, self.последняяКолонка + 1):
            colLetter = get_column_letter(stN)
            СтилиКолонки = self.sheet.column_dimensions[colLetter]
            for attribut in vars(СтилиКолонки):
                setattr(targetSheet.column_dimensions[colLetter], attribut,getattr(СтилиКолонки,attribut ))
        #Перенос объединения ячеек
        merged_cells__ranges = self.sheet.merged_cells.ranges
        for cellrange in merged_cells__ranges:
            if cellrange.max_row > max(номераСтрокТитула):
                continue
            targetSheet.merged_cells.add(copy.copy(cellrange))









