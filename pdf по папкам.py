#!/usr/bin/env python
# coding: utf-8

# In[4]:


# -*- coding: utf-8 -*-
#############################################################################################
#############################################################################################
#                    Скрипт читает файлы в текущей директории. Ищет названия в таблице.
#                    Если находит - создает папки с иерархией из таблицы
#
#############################################################################################
#############################################################################################


#############################################################################################
# Подключение модулей
#############################################################################################
import sys
import officeauto3_2
import copy
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
#import pandas as pd
import officeauto3_2
import re
import glob
import shutil
import sys
import copy
import PySimpleGUI as sg


#############################################################################################
# Настройки
#############################################################################################
имяВходногоФайла = './livebuildings.xlsx'
имяВыходногоФайла = ''
номерПервойСтроки = 1
номерПоследнейСтроки = 100000  # последняя строка
# адрес - C
#  округ - A
#  район - B


#############################################################################################
# Обработка таблицы
#############################################################################################
def обработкаТаблицы(имяВходногоФайла, имяВыходногоФайла, номерПервойСтроки, номерПоследнейСтроки):


    Книга1 = load_workbook(имяВходногоФайла, data_only=False, ) # Нам нужны не только данные!!!
    Лист1Книги1 = Книга1.worksheets[0]

    # Проверим папочку

    словарьАдресов ={}
    for номерСтроки in range(номерПервойСтроки, номерПоследнейСтроки +1):

        ИмяФайла = officeauto3_2.value(Лист1Книги1[f'C{номерСтроки}'])
        ИмяФайла = ИмяФайла.replace('/', '-')
        #ИмяФайла = ИмяФайла.replace('\\', '.')

        словарьАдресов[ИмяФайла] = {'округ': officeauto3_2.value(Лист1Книги1[f'A{номерСтроки}']),
                                    'район': officeauto3_2.value(Лист1Книги1[f'B{номерСтроки}'])}
        # Файл excel нам больше не нужен

    # Пройдемся по файлам
    #rp=input('ведите рабочую папку:')
    os.chdir(rp)
    def change_resolution(folder,old_resolution,new_resolution):
        files = os.listdir(folder)
        for file in files:
            if file.endswith(old_resolution):
                old_file_path = os.path.join(folder,file)
                new_file_path = os.path.join(folder, file.replace(old_resolution, new_resolution))
                os.rename (old_file_path, new_file_path)

    old_resolution = '.pdf'
    new_resolution = '.PDF'
    change_resolution(rp, old_resolution, new_resolution)

    for file in glob.glob("*.PDF"):

        ИмяФайла = file.split('.PDF')[0]
        ИмяФайла = ИмяФайла.replace('/', '-')
        #ИмяФайла = ИмяФайла.replace('.pdf', '.PDF')
        if not ИмяФайла in словарьАдресов:
            print(f'не найден: {ИмяФайла}')
            continue
        округ = f'./{ словарьАдресов[ИмяФайла]["округ"] }'
        район = f'./{ словарьАдресов[ИмяФайла]["район"] }'
        if not os.path.exists(округ):
            os.mkdir(f'./{округ}')
        if not os.path.exists(f'./{округ}/{район}'):
            os.mkdir(f'./{округ}/{район}')
        shutil.move(file, f'./{округ}/{район}/{file}')


# Задаем значения по умолчанию из блока "Настройки"
default_input_file = 'Ваш реестр.xlsx'
default_input_folder = 'rp'
email = 'Kovalev_A_D@moek.ru'
layout = [
    [sg.Text('В реестре колонки должны быть: Округ, Район, Адрес')],
    [[sg.Text("папка в которой будем работать: "), sg.Input(key="rp" ,change_submits=True), sg.FolderBrowse()]],
    [sg.Text('Имя входного файла:'), sg.Input(key='имяВходногоФайла', default_text=default_input_file), sg.FileBrowse()],

    [sg.Button('Запустить скрипт')],
    [sg.Text('Комментарии и предложения по работе скрипта направлять:'), sg.Text(email, text_color='blue', enable_events=True, key='-EMAIL-')]
]

window = sg.Window('Раскидаем файлы по Округам и районам согласно реестра', layout)

while True:
    event, values = window.read()

    if event == sg.WINDOW_CLOSED:
        break

    if event == "-EMAIL-":
        sg.popup('пишите письма')

    if event == 'Запустить скрипт':
        имяВходногоФайла = values['имяВходногоФайла']
        rp = values['rp']

        result = обработкаТаблицы(имяВходногоФайла, имяВыходногоФайла, номерПервойСтроки, номерПоследнейСтроки)

window.close()



#if __name__ == '__main__':
    #result = обработкаТаблицы(имяВходногоФайла, имяВыходногоФайла, номерПервойСтроки, номерПоследнейСтроки)

